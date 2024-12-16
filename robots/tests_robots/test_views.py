import json
import io

from django.test import TestCase, Client
from django.utils import timezone
from django.urls import reverse
from openpyxl import load_workbook

from robots.models import Robot


VALID_DATA = {
    "model": "R2",
    "version": "D2",
    "created": "2022-12-31 23:59:59"
    }


class RobotCreationAPITests(TestCase):
    def setUp(self):
        self.client = Client()
        self.url = reverse('robot_creation_API')
        self.valid_data = VALID_DATA

    def test_robot_creation_with_valid_data(self):
        response = self.client.post(
            self.url,
            json.dumps(self.valid_data),
            content_type='application/json'
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response.json()['Результат'],
            'Данные о производстве робота модели R2, версии D2 приняты'
        )
        self.assertEqual(Robot.objects.count(), 1)

        robot = Robot.objects.first()
        self.assertEqual(robot.model, 'R2')
        self.assertEqual(robot.version, 'D2')
        self.assertEqual(robot.serial, 'R2-D2')

    def test_robot_creation_with_forbidden_model(self):
        invalid_data = {
            "model": "D5",
            "version": "D2",
            "created": "2022-12-31 23:59:59"
            }
        response = self.client.post(
            self.url,
            json.dumps(invalid_data),
            content_type='application/json'
        )
        self.assertEqual(response.status_code, 400)
        self.assertIn('error', response.json())
        self.assertIn(
            'Эта модель робота не разрешена',
            response.json()['error']['__all__']
        )

    def test_robot_creation_with_invalid_version(self):
        invalid_data = {
            "model": "R2",
            "version": "Invalid",
            "created": "2022-12-31 23:59:59"
            }
        response = self.client.post(
            self.url,
            json.dumps(invalid_data),
            content_type='application/json'
        )
        self.assertEqual(response.status_code, 400)
        self.assertIn('error', response.json())

    def test_robot_creation_with_invalid_json(self):
        response = self.client.post(
            self.url,
            data='',
            content_type='application/json'
            )
        self.assertEqual(response.status_code, 400)
        self.assertIn('error', response.json())
        self.assertEqual(
            response.json()['error'],
            'Некорректный формат, передайте данные в формате JSON.'
            )


class JsonRobotCreationViewTests(TestCase):
    def setUp(self):
        self.client = Client()
        self.url = reverse('json_robot_creation')
        self.valid_data = VALID_DATA

    def test_robot_creation_with_valid_data(self):
        response = self.client.post(
            self.url,
            {'json_data': json.dumps(self.valid_data)}
        )

        self.assertEqual(response.status_code, 200)
        self.assertIn('result_message', response.context)

        self.assertEqual(Robot.objects.count(), 1)
        robot = Robot.objects.first()
        self.assertEqual(robot.model, 'R2')
        self.assertEqual(robot.version, 'D2')
        self.assertEqual(robot.serial, 'R2-D2')

    def test_robot_creation_with_invalid_json(self):
        response = self.client.post(
            self.url,
            {'json_data': "Invalid JSON"}
        )
        self.assertEqual(response.status_code, 400)
        self.assertEqual(response.context, None)

    def test_robot_creation_with_invalid_model(self):
        invalid_data = {
            "model": "Invalid",
            "version": "D2",
            "created": "2022-12-31 23:59:59"
            }
        response = self.client.post(
            self.url,
            {'json_data': json.dumps(invalid_data)}
        )
        self.assertFalse(response.context['form'].is_valid())
        self.assertEqual(Robot.objects.count(), 0)

    def test_robot_creation_with_frobidden_model(self):
        invalid_data = {
            "model": "R25",
            "version": "D2",
            "created": "2022-12-31 23:59:59"
            }
        response = self.client.post(
            self.url,
            {'json_data': json.dumps(invalid_data)}
        )
        self.assertFalse(response.context['form'].is_valid())
        self.assertEqual(Robot.objects.count(), 0)

    def test_robot_creation_with_invalid_verision(self):
        invalid_data = {
            "model": "R2",
            "version": "D25555",
            "created": "2022-12-31 23:59:59"
            }
        response = self.client.post(
            self.url,
            {'json_data': json.dumps(invalid_data)}
        )
        self.assertFalse(response.context['form'].is_valid())
        self.assertEqual(Robot.objects.count(), 0)

    def test_robot_creation_without_verision(self):
        invalid_data = {
            "model": "R2",
            "created": "2022-12-31 23:59:59"
            }
        response = self.client.post(
            self.url,
            {'json_data': json.dumps(invalid_data)}
        )
        self.assertFalse(response.context['form'].is_valid())
        self.assertEqual(Robot.objects.count(), 0)


class ProductReportViewTests(TestCase):

    def setUp(self):
        self.client = Client()
        self.url = reverse('production_report')

    def create_robot(self, model, version, created):
        """Вспомогательный метод для создания тестового робота."""
        return Robot.objects.create(
            model=model, version=version, created=created
            )

    def test_product_report(self):
        # Создаем роботов
        self.create_robot(model='R2', version='D2', created=timezone.now())
        self.create_robot(model='X5', version='LT', created=timezone.now())
        self.create_robot(model='R2', version='D3', created=timezone.now())

        response = self.client.get(reverse('production_report'))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response['Content-Disposition'],
            'attachment; filename=proudction_report.xlsx'
            )

        excel_file = io.BytesIO(response.content)
        workbook = load_workbook(excel_file)

        # Проверяем наличие листов
        self.assertEqual(len(workbook.sheetnames), 2)

        # Проверяем корректность заполнения листов
        sheet_a = workbook['R2']
        self.assertEqual(sheet_a.cell(row=1, column=1).value, 'Модель')
        self.assertEqual(sheet_a.cell(row=2, column=1).value, 'R2')
        self.assertEqual(sheet_a.cell(row=2, column=2).value, 'D2')
        self.assertEqual(sheet_a.cell(row=2, column=3).value, 1)
        self.assertEqual(sheet_a.cell(row=3, column=1).value, 'R2')
        self.assertEqual(sheet_a.cell(row=3, column=2).value, 'D3')
        self.assertEqual(sheet_a.cell(row=3, column=3).value, 1)

        sheet_b = workbook['X5']
        self.assertEqual(sheet_b.cell(row=1, column=1).value, 'Модель')
        self.assertEqual(sheet_b.cell(row=2, column=1).value, 'X5')
        self.assertEqual(sheet_b.cell(row=2, column=2).value, 'LT')
        self.assertEqual(sheet_b.cell(row=2, column=3).value, 1)

    def test_production_report_no_data(self):

        response = self.client.get(reverse('production_report'))
        # Проверяем статус ответа
        self.assertEqual(response.status_code, 400)
